import os
import time 
import keyring
from cryptography.fernet import Fernet
from dotenv import load_dotenv
import tkinter as tk
from tkinter import simpledialog, messagebox
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.edge.options import Options
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException
from requests.exceptions import ConnectionError
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
import shutil
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

#INSTALADORES
#pip install msedge-selenium-tools
#pip install webdriver-manager


from subprocess import CREATE_NO_WINDOW

def solicitar_credenciales():
    def confirmar():
        cuenta = cuenta_entry.get().strip()
        contrasena = contrasena_entry.get().strip()
        correo = correo_entry.get().strip()
        contrasena_aplicacion = contrasena_app_entry.get().strip()

        if not cuenta:
            messagebox.showerror("Error", "Debe ingresar una cuenta.")
            return
        if not contrasena:
            messagebox.showerror("Error", "Debe ingresar una contraseña.")
            return
        if not correo:
            messagebox.showerror("Error", "Debe ingresar un correo electrónico.")
            return
        if not contrasena_aplicacion:
            messagebox.showerror("Error", "Debe ingresar una contraseña de aplicación.")
            return

        root.quit()  # Salir del bucle principal
        root.destroy()  # Cerrar la ventana
        nonlocal result
        result = (cuenta, contrasena, correo, contrasena_aplicacion)

    # Configuración inicial
    root = tk.Tk()
    root.title("Iniciar Sesión")
    root.configure(bg="#4B0082") 
    root.geometry("400x400")
    root.resizable(False, False)

    # Títulos y estilos
    tk.Label(
        root, text="Iniciar Sesión", font=("Arial", 18, "bold"), bg="#4B0082", fg="white"
    ).pack(pady=10)

    # Campo de entrada para la cuenta
    tk.Label(root, text="Cuenta(Jira):", font=("Arial", 12), bg="#4B0082", fg="white").pack(
        pady=5
    )
    cuenta_entry = tk.Entry(root, font=("Arial", 12), width=30)
    cuenta_entry.pack(pady=5)

    # Campo de entrada para la contraseña
    tk.Label(
        root, text="Contraseña(Jira):", font=("Arial", 12), bg="#4B0082", fg="white"
    ).pack(pady=5)
    contrasena_entry = tk.Entry(root, font=("Arial", 12), width=30, show="*")
    contrasena_entry.pack(pady=5)

    # Campo de entrada para el correo electrónico
    tk.Label(
        root, text="Correo Electrónico:", font=("Arial", 12), bg="#4B0082", fg="white"
    ).pack(pady=5)
    correo_entry = tk.Entry(root, font=("Arial", 12), width=30)
    correo_entry.pack(pady=5)

    # Campo de entrada para la contraseña de aplicación
    tk.Label(
        root, text="Contraseña de Aplicación:", font=("Arial", 12), bg="#4B0082", fg="white"
    ).pack(pady=5)
    contrasena_app_entry = tk.Entry(root, font=("Arial", 12), width=30, show="*")
    contrasena_app_entry.pack(pady=5)

    # Botón para confirmar
    confirmar_btn = tk.Button(
        root,
        text="Confirmar",
        font=("Arial", 12),
        bg="white",
        fg="#4B0082",
        command=confirmar,
    )
    confirmar_btn.pack(pady=20)

    result = None
    root.mainloop()  # Mostrar la ventana
    return result




def guardar_clave_en_credenciales(nombre, valor):
    
    try:
        keyring.set_password("my_application", nombre, valor) 
        print(f"Clave {nombre} guardada/actualizada en el Windows Credential Manager.")
    except Exception as e:
        print(f"Error al guardar la clave en el Credential Manager: {e}")

def obtener_clave_desde_credenciales(nombre):
    
    try:
        valor = keyring.get_password("my_application", nombre)
        if valor:
            return valor
        else:
            print(f"No se encontró la clave {nombre} en el Credential Manager.")
            return None
    except Exception as e:
        print(f"Error al obtener la clave desde el Credential Manager: {e}")
        return None


def generar_clave():
    nueva_clave = Fernet.generate_key().decode()  
    guardar_clave_en_credenciales("ENCRYPTION_KEY", nueva_clave)  
    return nueva_clave

def validar_crear_carpeta_input():
    carpeta_input = './Input'
    if not os.path.exists(carpeta_input):
        os.makedirs(carpeta_input)
        print("La carpeta 'Input' no existía, se ha creado correctamente.")
    else:
        print("La carpeta 'Input' ya existe.")

def guardar_credenciales(cuenta, contrasena, correo, contrasena_aplicacion):

    validar_crear_carpeta_input()

    if not cuenta or not contrasena or not correo or not contrasena_aplicacion:
        print("No se ingresaron credenciales válidas.")
        return

    clave = generar_clave()
    cipher = Fernet(clave.encode())

    cuenta_encriptada = cipher.encrypt(cuenta.encode()).decode()
    contrasena_encriptada = cipher.encrypt(contrasena.encode()).decode()
    correo_encriptado = cipher.encrypt(correo.encode()).decode()
    contrasena_aplicacion_encriptada = cipher.encrypt(contrasena_aplicacion.encode()).decode()

    with open("Input/.env", "w") as archivo:
        archivo.write(f"CUENTA={cuenta_encriptada}\n")
        archivo.write(f"CONTRASENA={contrasena_encriptada}\n")
        archivo.write(f"CORREO={correo_encriptado}\n")
        archivo.write(f"CONTRASENA_APLICACION={contrasena_aplicacion_encriptada}\n")

    print("Credenciales encriptadas y guardadas correctamente en .env.")
    print(f"Clave actualizada en el Credential Manager: ENCRYPTION_KEY={clave}")

def cargar_credenciales():
    load_dotenv('./Input/.env')

    clave = obtener_clave_desde_credenciales("ENCRYPTION_KEY")
    if not clave:
        print("Error: La clave de desencriptado no está configurada en el Windows Credential Manager.")
        return

    cipher = Fernet(clave.encode())

    cuenta_encriptada = os.getenv("CUENTA")
    contrasena_encriptada = os.getenv("CONTRASENA")
    correo_encriptado = os.getenv("CORREO")
    contrasena_aplicacion_encriptada = os.getenv("CONTRASENA_APLICACION")

    if not cuenta_encriptada or not contrasena_encriptada or not correo_encriptado or not contrasena_aplicacion_encriptada:
        print("Error: No se encontraron credenciales en el archivo .env.")
        return
    
    cuenta = cipher.decrypt(cuenta_encriptada.encode()).decode()
    contrasena = cipher.decrypt(contrasena_encriptada.encode()).decode()
    correo = cipher.decrypt(correo_encriptado.encode()).decode()
    contrasena_aplicacion = cipher.decrypt(contrasena_aplicacion_encriptada.encode()).decode()

    #print("Credenciales cargadas correctamente:")
    #print(f"Cuenta: {cuenta}")
    #print(f"Contraseña: {contrasena}")
    #print(f"Correo: {correo}")
    #print(f"Contraseña de Aplicación: {contrasena_aplicacion}")

    return cuenta, contrasena,correo,contrasena_aplicacion
def  configuracion_navegador(cwd):
    options = Options()
    prefs = {
        "download.default_directory": cwd,  
        "safebrowsing.enabled": "false",    
        "profile.default_content_settings.popups": 0,
    }
    options.add_experimental_option("prefs", prefs)
    options.add_argument("--disable-breakpad") 
    options.add_argument("disable-infobars")    
    options.add_argument("--disable-sync")
    options.add_argument("--log-level=3")  
    options.add_argument("--inprivate")


    
    #options.add_argument("--headless") #Descomentar codigo al final 🔥
    return options
def instancia_webdriver_edge(options, url):
    try:
        service = Service(EdgeChromiumDriverManager().install())
        service.creationflags = getattr(os, "CREATE_NO_WINDOW", 0)
        driver = webdriver.Edge(service=service, options=options)
        #driver.minimize_window()  # Minimiza la ventana al iniciar
        driver.get(url)   
        
        return driver
    except ConnectionError:
        messagebox.showerror("Error", "No se pudo instalar el driver!\nRevisa tu conexión ⚠️")
        return None
        
    

def autenticacion_itsm(driver,cuenta,contrasena):
    print('Prueba...')
    try:
        if(driver.find_element(By.XPATH,'//*[@id="main-message"]/h1/span')):
            messagebox.showerror('No estas conectado a internet!\n Revise su conexion a internet⚠️')
            return False
    except NoSuchElementException:
        pass
    try:
        wait = WebDriverWait(driver ,20)
        print('Entre')
        btn_auten = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="microsoft-auth-button"]')))
        btn_auten.click()
        print('Entre2')
        input_cuenta= wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="i0116"]')))
        input_cuenta.send_keys(cuenta)
        btn_siguiente = wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="idSIButton9"]')))
        btn_siguiente.click()
        input_contra = wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="i0118"]')))
        input_contra.send_keys(contrasena)
        
        btn_iniciar = wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="idSIButton9"]')))
        btn_iniciar.click()
        
        btn_no_iniciada = wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="idBtn_Back"]')))
        btn_no_iniciada.click()
        print('Aqui')
        
        try:
            print('todo good')
            btn_continuar = wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="login-submit"]')))
            btn_continuar.click()
            time.sleep(20)
            print('mini siesta terminada')
            return driver
        except NoSuchElementException:
            print('No encontre 😠')
            pass

        return driver
    except TimeoutException as e:
        print(f"Error: Timeout esperando un elemento. Detalles: {str(e)}")
        return False
def navegacion_itsm(driver):
    print('a')
    url_proyecto = 'https://servicios-it-corfi.atlassian.net/jira/servicedesk/projects/MS/queues/custom/50'
    wait = WebDriverWait(driver,10)
    driver.get(url_proyecto)
    time.sleep(10)
    url_filtros = 'https://servicios-it-corfi.atlassian.net/jira/filters'
    driver.get(url_filtros)
    input_filtro = wait.until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/div[1]/div[2]/div[4]/div[1]/main/div/div/div[1]/div/div[1]/div[2]/div/div[1]/div/div/input')))
    input_filtro.send_keys('Todos ABIERTOS NIVEL 2')
    time.sleep(10)
    click_filtro = wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="ak-main-content"]/div/div/div[1]/div/div[2]/div[2]/table/tbody/tr/td[2]/div/a')))
    click_filtro.click()
    btn_excel = wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="com.atlassian.jira.spreadsheets__open-in-excel"]/span')))
    btn_excel.click()
    print(f"URL de la nueva pestaña: {driver.current_url}")
    pestana_original = driver.current_window_handle
    print('descargando')
    WebDriverWait(driver, 10).until(lambda d: len(d.window_handles) > 1)
    new_tab = [tab for tab in driver.window_handles if tab != pestana_original][0]
    driver.switch_to.window(new_tab)
    
    print(f"URL de la nueva pestaña: {driver.current_url}")


    time.sleep(10)
    btn_excel_desktop = wait.until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[2]/div/div/div[2]/div/div/section/div[3]/a[1]/span')))
    btn_excel_desktop.click()
   

    time.sleep(10)
    print('termine')
    return driver

def renombrar_excel():
    for x in os.listdir('Input'):
        if 'jira-search' in x :
            try:
                os.rename(f'Input/{x}','Input/FiltroCNI.xlsx')
            except:
                os.remove('Input/FiltroCNI.xlsx')
                os.rename(f'Input/{x}','Input/FiltroCNI.xlsx')

# pip install openpyxl 
def manipular_excel_y_cargar_sharepoint(driver):
    
    wb = load_workbook('Input/FiltroCNI.xlsx')
    sheet = wb.active

    data=[]

    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    df = pd.DataFrame(data)

    df.columns = df.iloc[0]
    df = df[1:]

    # Filtrar las filas por "Categoría de estado"
    filtrados = df[df['Categoría de estado'].isin(['Asignación & Análisis'])] 

            # Crear la carpeta de salida si no existe
    carpeta_output = './Output'
    if not os.path.exists(carpeta_output):
        os.makedirs(carpeta_output)

    output_path = 'Output/Datos_Filtrados_CNI.xlsx'

    if not filtrados.empty:
        print('Se encontraron filas con "Asignación & Análisis" en la columna "Categoría de estado".')

        # Guardar los datos filtrados en un nuevo archivo Excel
        filtrados.to_excel(output_path, index=False)

        print(f"\nLos datos filtrados se han guardado en: {output_path}")

    else:
         # Crear un archivo Excel vacío si no se encuentran datos
        df_vacio = pd.DataFrame(columns=df.columns)  # Crear un DataFrame vacío con las mismas columnas
        df_vacio.to_excel(output_path, index=False)  # Guardar el archivo vacío
        print('No se encontraron filas con "Asignación & Análisis" en la columna "Categoría de estado".')



def asignar_correo(destinatario, asunto, mensaje,cuenta,contrasena):
    SMTP_SERVER = "smtp.office365.com"
    SMTP_PORT = 587
    try:
        msg = MIMEMultipart()
        msg['From'] = cuenta
        msg['To'] = destinatario
        msg['Subject'] = asunto

        msg.attach(mensaje)

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(cuenta, contrasena)
            server.sendmail(cuenta, destinatario, msg.as_string())
        
        print("Correo enviado con éxito.")
    except Exception as e:
        print(f"Error al enviar el correo: {e}")

def enviar_correo(cuenta,contrasena,correo_equipo_teams):

    caso_nuevo = pd.read_excel('Output/Datos_Filtrados_CNI.xlsx')
    correo_asignado = pd.read_excel('Input/Persona asignada CNI.xlsx')

    mensajes_acumulados = []

    for i, row in correo_asignado.iterrows():

        destinatario = row.get('correo')
        persona_asignada = row.get('Persona asignada')

        if pd.notna(destinatario):

            for j, caso_row in caso_nuevo.iterrows():
                # Extraer los datos del caso
                clave = caso_row['Clave']
                creada = caso_row['Creada']
                prioridad = caso_row['Prioridad']
                resumen = caso_row['Resumen']
                p_encargada = caso_row['Persona asignada']
                categoría_estado = caso_row['Categoría de estado']

                mensaje_correo = MIMEText(f"""
                                <html>
                                <head>
                                    <style>
                                    body {{
                                        font-family: Arial, sans-serif;
                                        font-size: 14px;
                                        color: #333333;
                                    }}
                                    p {{
                                        margin: 10px 0;
                                    }}
                                    .resumen {{
                                        color: #0056b3;
                                        font-weight: bold;
                                    }}
                                    .fecha {{
                                        color: #888888;
                                    }}
                                    </style>
                                </head>
                                <body>
                                    <p>Buen día <strong>{persona_asignada},</strong></p>
                                    <p>Acaba de llegar un nuevo caso: <span class="resumen">{clave}</span> "{resumen}" y se encuentra en nivel de prioridad: 
                                    <em>"{prioridad}"</em> con categoria de estado <em>{categoría_estado}</em>  y fue creado el; <span class="fecha">{creada}</span>.</p>
                                    <p>Por favor, dar primera respuesta.</p>
                                    <p>¡Gracias!</p>
                                </body>
                                </html>
                            """, 'html')

                asunto=f'Caso {clave} pendiente de respuesta'

                asignar_correo(destinatario, asunto, mensaje_correo,cuenta,contrasena)
                print(f'Correo enviado con exito a {destinatario}')
                # envio de mensaje a grupo de teams por medio de correo de teams
                mensaje_individual=f"""
                <p>---------------------------------------------------------------------------------------------------</p>
                <p>✔ Buen día <strong>{persona_asignada},</strong></p>
                <p>Acaba de llegar un nuevo caso: <span class="resumen">{clave}</span> "{resumen}" y se encuentra en nivel de prioridad: 
                <em>"{prioridad}"</em> con categoria de estado <strong>{categoría_estado}</strong>  y fue creado el; <strong><span class="fecha">{creada}</span></strong>.</p>
                """

                # Añadir el mensaje a la lista de mensajes 
                mensajes_acumulados.append(mensaje_individual)

                # Consolidar todos los mensajes en un solo correo para el equipo de Teams
        if mensajes_acumulados:
            mensajes_html = ''.join(mensajes_acumulados)
            mensaje_equipo = MIMEText(f"""
                <html>
                <head></head>
                <body>
                    {mensajes_html}
                    <p>---------------------------------------------------------------------------------------------------</p>
                    <p>Por favor, revisar estos casos nuevos.</p>
                    <p>¡Gracias!</p>
                </body>
                </html>
            """, 'html')
            asunto_equipo = "Resumen de casos nuevos"

            asignar_correo(correo_equipo_teams, asunto_equipo, mensaje_equipo,cuenta,contrasena)
        else:
            print('No hay casos nuevos')


def main ():   

    if not os.path.exists("Input/.env"):
        cuenta, contrasena,correoelectronico,contrasena_de_aplicacion = solicitar_credenciales()
        if cuenta and contrasena:
            guardar_credenciales(cuenta, contrasena,correoelectronico,contrasena_de_aplicacion)
    else:
        cuenta , contrasena,correoelectronico,contrasena_de_aplicacion = cargar_credenciales()
        print('Nice')
    
    cwd = os.path.abspath('Input')
    
    options = configuracion_navegador(cwd=cwd)
    url_jira = "https://id.atlassian.com/login"
    
    driver = instancia_webdriver_edge(options=options,url= url_jira)
    print(f'Driver: {driver}')
    driver = autenticacion_itsm(driver=driver,cuenta=cuenta,contrasena=contrasena)
    navegacion_itsm(driver=driver)
    renombrar_excel()
    manipular_excel_y_cargar_sharepoint(driver)
    correo_equipo_teams='cc9f3499.axity.com@amer.teams.ms'#Correo de equipo de teams
    enviar_correo(correoelectronico,contrasena_de_aplicacion,correo_equipo_teams)


main()
